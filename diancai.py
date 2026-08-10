from datetime import datetime, timedelta
import os
import random
import shutil
import sqlite3
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from PIL import Image, ImageDraw, ImageFont, ImageTk

# --- 动态获取当前路径（完美适配 Python 运行及 PyInstaller 打包后的 exe） ---
if getattr(sys, "frozen", False):
  BASE_DIR = os.path.dirname(sys.executable)
else:
  BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DB_NAME = os.path.join(BASE_DIR, "restaurant.db")
IMAGE_DIR = os.path.join(BASE_DIR, "image")


# --- 核心路径处理工具函数（解决绝对路径硬编码问题） ---
def get_relative_path(abs_path):
  """将绝对路径转为相对于程序根目录(BASE_DIR)的相对路径，确保数据库跨环境可移植"""
  if not abs_path:
    return ""
  if not os.path.isabs(abs_path):
    return abs_path
  try:
    rel = os.path.relpath(abs_path, BASE_DIR)
    if rel.startswith(".."):
      return abs_path  # 如果不在项目目录下，则保留原样
    return rel
  except Exception:
    return abs_path


def resolve_img_path(path):
  """根据当前运行环境动态解析图片路径（兼容旧绝对路径、跨系统及相对路径）"""
  if not path:
    return ""
  if os.path.isabs(path) and os.path.exists(path):
    return path

  normalized = path.replace("\\", "/")
  if "/image/" in normalized:
    sub_part = normalized.split("/image/")[-1]
    candidate = os.path.join(IMAGE_DIR, sub_part.replace("/", os.sep))
    if os.path.exists(candidate):
      return candidate

  return os.path.join(BASE_DIR, path)


def init_db():
  conn = sqlite3.connect(DB_NAME)
  try:
    cursor = conn.cursor()
    # 菜品表
    cursor.execute("""
            CREATE TABLE IF NOT EXISTS dishes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                price REAL,
                category TEXT,
                image_path TEXT
            )
        """)
    # 交易流水表（后台导入用）
    cursor.execute("""
            CREATE TABLE IF NOT EXISTS transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                trans_time TEXT,
                amount REAL
            )
        """)
    # 已提交订单主表
    cursor.execute("""
            CREATE TABLE IF NOT EXISTS orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_time TEXT,
                items_detail TEXT,
                total_amount REAL,
                trans_id INTEGER
            )
        """)
    # 订单明细子表
    cursor.execute("""
            CREATE TABLE IF NOT EXISTS order_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id INTEGER,
                name TEXT,
                price REAL,
                count INTEGER,
                image_path TEXT
            )
        """)
    # 软件加载时清空临时导入内容
    cursor.execute("DELETE FROM transactions")
    conn.commit()
  finally:
    conn.close()


def auto_sync_images_on_start():
  """软件启动时自动扫描分类文件夹，并同步菜品（仅在数据库为空时或初始化）"""
  if not os.path.exists(IMAGE_DIR):
    return
  conn = sqlite3.connect(DB_NAME)
  try:
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM dishes")
    if cursor.fetchone()[0] == 0:
      for category_name in os.listdir(IMAGE_DIR):
        category_path = os.path.join(IMAGE_DIR, category_name)
        if os.path.isdir(category_path):
          for filename in os.listdir(category_path):
            if filename.lower().endswith((".png", ".jpg", ".jpeg", ".webp")):
              base_name = os.path.splitext(filename)[0]
              parts = base_name.split("_")
              name = parts[0]
              price = 0.0
              if len(parts) > 1:
                try:
                  price = float(parts[1])
                except ValueError:
                  pass
              abs_img_path = os.path.join(category_path, filename)
              rel_img_path = get_relative_path(abs_img_path)
              cursor.execute(
                  "INSERT INTO dishes (name, price, category, image_path) VALUES"
                  " (?, ?, ?, ?)",
                  (name, price, category_name, rel_img_path),
              )
      conn.commit()
  except Exception as e:
    print(f"启动自动同步图片异常: {e}")
  finally:
    conn.close()


init_db()
auto_sync_images_on_start()


def get_pil_font(size):
  """跨平台获取支持中文的字体，优化 Windows 打包兼容性"""
  font_paths = []
  if sys.platform.startswith("win"):
    font_paths = [
        "C:/Windows/Fonts/msyh.ttc",
        "C:/Windows/Fonts/simhei.ttf",
        "C:/Windows/Fonts/simsun.ttc",
        "C:/Windows/Fonts/arial.ttf",
    ]
  elif sys.platform.startswith("darwin"):
    font_paths = ["/System/Library/Fonts/PingFang.ttc", "/Library/Fonts/Arial.ttf"]
  else:
    font_paths = [
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
        "/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
  for path in font_paths:
    if os.path.exists(path):
      try:
        return ImageFont.truetype(path, size)
      except Exception:
        continue
  try:
    return ImageFont.load_default()
  except Exception:
    return None


class RestaurantApp(tk.Tk):

  def __init__(self):
    super().__init__()
    self.title("点菜收银系统")
    self.geometry("1350x750")

    nav_frame = tk.Frame(self, bg="#f0f0f0", height=45)
    nav_frame.pack(side=tk.TOP, fill=tk.X)

    tk.Button(
        nav_frame,
        text=" 🍽️ 前台点餐与订单管理 ",
        font=("Arial", 11, "bold"),
        command=self.show_page1,
    ).pack(side=tk.LEFT, padx=10, pady=5)
    tk.Button(
        nav_frame,
        text=" ⚙️ 后台管理中心 ",
        font=("Arial", 11, "bold"),
        command=self.show_page2,
    ).pack(side=tk.LEFT, padx=10, pady=5)

    self.container = tk.Frame(self)
    self.container.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
    self.container.rowconfigure(0, weight=1)
    self.container.columnconfigure(0, weight=1)

    self.pages = {}
    self.pages["Page1"] = PageOrdering(self.container)
    self.pages["Page2"] = PageAdmin(self.container, self.pages["Page1"])

    self.show_page1()

  def show_page1(self):
    self.pages["Page1"].refresh_data()
    self.pages["Page1"].tkraise()

  def show_page2(self):
    self.pages["Page2"].refresh_data()
    self.pages["Page2"].tkraise()


# --- 第一页：前台点餐与三栏布局 ---
class PageOrdering(tk.Frame):

  def __init__(self, parent):
    super().__init__(parent)
    self.grid(row=0, column=0, sticky="nsew")
    self.image_cache = {}  # Tkinter PhotoImage 持久化缓存
    self.pil_image_cache = {}  # 原始 PIL 缩放图缓存，彻底消除重复读盘卡顿
    self.cart = {}
    self.cart_ui_elements = {}
    self.editing_order_id = None
    self.is_initialized = False

    self.left_pane = tk.Frame(self)
    self.left_pane.pack(
        side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 5), pady=5
    )

    self.middle_pane = ttk.LabelFrame(
        self, text=" 🛒 当前点菜详情与购物车 ", width=320, padding=10
    )
    self.middle_pane.pack(
        side=tk.LEFT, fill=tk.BOTH, expand=False, padx=5, pady=5
    )
    self.middle_pane.pack_propagate(False)

    self.right_pane = ttk.LabelFrame(
        self, text=" 📋 已提交历史订单列表 ", width=360, padding=10
    )
    self.right_pane.pack(
        side=tk.LEFT, fill=tk.BOTH, expand=False, padx=(5, 10), pady=5
    )
    self.right_pane.pack_propagate(False)

    tk.Label(
        self.left_pane,
        text="【前台点餐区 - 菜品分类】",
        font=("Arial", 14, "bold"),
    ).pack(anchor="w", pady=2)
    self.notebook = ttk.Notebook(self.left_pane)
    self.notebook.pack(fill=tk.BOTH, expand=True, pady=2)

    self.setup_cart_ui()
    self.setup_order_history_ui()

  def setup_cart_ui(self):
    cart_container = tk.Frame(self.middle_pane)
    cart_container.pack(fill=tk.BOTH, expand=True, pady=5)

    self.cart_canvas = tk.Canvas(cart_container, bg="white", highlightthickness=0)
    cart_scrollbar = ttk.Scrollbar(
        cart_container, orient="vertical", command=self.cart_canvas.yview
    )
    self.cart_items_frame = ttk.Frame(self.cart_canvas)

    self.cart_items_frame.bind(
        "<Configure>",
        lambda e: self.cart_canvas.configure(
            scrollregion=self.cart_canvas.bbox("all")
        ),
    )
    self.cart_canvas.create_window(
        (0, 0), window=self.cart_items_frame, anchor="nw"
    )
    self.cart_canvas.configure(yscrollcommand=cart_scrollbar.set)

    self.cart_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    cart_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    bottom_frame = tk.Frame(self.middle_pane)
    bottom_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=5)

    self.status_frame = tk.Frame(bottom_frame)
    self.status_frame.pack(fill=tk.X, pady=2)
    self.edit_status_label = tk.Label(
        self.status_frame, text="", fg="#FF9800", font=("Arial", 9, "bold")
    )
    self.edit_status_label.pack(side=tk.LEFT)

    self.cancel_edit_btn = tk.Button(
        self.status_frame,
        text="取消编辑",
        command=self.cancel_edit,
        bg="#e0e0e0",
        font=("Arial", 8),
    )

    actions_frame = tk.Frame(bottom_frame)
    actions_frame.pack(fill=tk.X, pady=2)

    self.total_label = tk.Label(
        actions_frame, text="总计: ¥0.00", font=("Arial", 12, "bold"), fg="#d32f2f"
    )
    self.total_label.pack(side=tk.LEFT, padx=2)

    tk.Button(
        actions_frame,
        text="清空",
        command=self.clear_cart,
        bg="#e0e0e0",
        font=("Arial", 9),
    ).pack(side=tk.RIGHT, padx=2)

    self.submit_btn = tk.Button(
        actions_frame,
        text="提交订单",
        command=self.submit_order,
        bg="#4CAF50",
        fg="white",
        font=("Arial", 10, "bold"),
    )
    self.submit_btn.pack(side=tk.RIGHT, padx=5)

  def setup_order_history_ui(self):
    history_container = tk.Frame(self.right_pane)
    history_container.pack(fill=tk.BOTH, expand=True, pady=5)

    self.order_tree = ttk.Treeview(
        history_container, columns=("时间", "菜品详情", "金额"), show="headings"
    )
    self.order_tree.heading("时间", text="下单时间")
    self.order_tree.heading("菜品详情", text="订单内容")
    self.order_tree.heading("金额", text="金额")

    self.order_tree.column("时间", width=110, anchor="center")
    self.order_tree.column("菜品详情", width=160, anchor="w")
    self.order_tree.column("金额", width=70, anchor="center")

    self.order_tree.bind("<<TreeviewSelect>>", self.on_order_select)

    order_scrollbar = ttk.Scrollbar(
        history_container, orient="vertical", command=self.order_tree.yview
    )
    self.order_tree.configure(yscrollcommand=order_scrollbar.set)

    self.order_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    order_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    btn_clear_history = tk.Button(
        self.right_pane,
        text="清空所有历史订单",
        command=self.clear_order_history,
        bg="#9e9e9e",
        fg="white",
        font=("Arial", 9),
    )
    btn_clear_history.pack(side=tk.BOTTOM, fill=tk.X, pady=2)

    btn_del_single = tk.Button(
        self.right_pane,
        text="删除选中历史订单",
        command=self.delete_selected_order,
        bg="#f44336",
        fg="white",
        font=("Arial", 9, "bold"),
    )
    btn_del_single.pack(side=tk.BOTTOM, fill=tk.X, pady=2)

  def on_order_select(self, event):
    selected_items = self.order_tree.selection()
    if not selected_items:
      return

    order_id = selected_items[0]
    self.editing_order_id = order_id

    self.cart.clear()
    conn = sqlite3.connect(DB_NAME)
    try:
      cursor = conn.cursor()
      cursor.execute(
          "SELECT name, price, count, image_path FROM order_items WHERE"
          " order_id = ?",
          (order_id,),
      )
      for row in cursor.fetchall():
        name, price, count, image_path = row
        self.cart[name] = {
            "price": price,
            "count": count,
            "image_path": image_path,
        }
    finally:
      conn.close()
    self.update_cart_ui_display()

  def cancel_edit(self):
    self.editing_order_id = None
    self.clear_cart()
    messagebox.showinfo("提示", "已退出订单修改模式")

  def delete_selected_order(self):
    selected_items = self.order_tree.selection()
    if not selected_items:
      messagebox.showwarning("提示", "请先在右侧列表中选中要删除的历史订单！")
      return

    order_id = selected_items[0]
    if messagebox.askyesno("确认删除", "确定要彻底删除该历史订单吗？"):
      conn = sqlite3.connect(DB_NAME)
      try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM orders WHERE id = ?", (order_id,))
        cursor.execute("DELETE FROM order_items WHERE order_id = ?", (order_id,))
        conn.commit()
      finally:
        conn.close()

      if self.editing_order_id == str(order_id):
        self.editing_order_id = None
        self.clear_cart()

      self.refresh_order_history()
      messagebox.showinfo("成功", "订单已成功删除！")

  def get_cached_image(self, cache_key, abs_img_path, size=(90, 90)):
    """高性能图像缓存，结合 PIL 缓存层避免反复读盘缩放引起卡顿"""
    if cache_key in self.image_cache:
      return self.image_cache[cache_key]

    pil_cache_key = f"{abs_img_path}_{size[0]}_{size[1]}"
    if pil_cache_key in self.pil_image_cache:
      base_img = self.pil_image_cache[pil_cache_key].copy()
    else:
      base_img = Image.new("RGB", size, (255, 255, 255))
      if abs_img_path and os.path.exists(abs_img_path):
        try:
          pil_img = Image.open(abs_img_path).convert("RGB")
          pil_img.thumbnail(size)
          w, h = pil_img.size
          base_img.paste(pil_img, ((size[0] - w) // 2, (size[1] - h) // 2))
        except Exception:
          pass
      self.pil_image_cache[pil_cache_key] = base_img.copy()

    tk_img = ImageTk.PhotoImage(base_img)
    self.image_cache[cache_key] = tk_img
    return tk_img

  def refresh_data(self):
    if not self.is_initialized:
      self.init_dish_ui()
      self.is_initialized = True
    else:
      self.refresh_order_history()

  def init_dish_ui(self):
    for tab in self.notebook.tabs():
      self.notebook.forget(tab)

    conn = sqlite3.connect(DB_NAME)
    try:
      cursor = conn.cursor()
      cursor.execute("SELECT DISTINCT category FROM dishes")
      categories = [row[0] for row in cursor.fetchall()]

      if not categories:
        empty_label = ttk.Label(
            self.notebook,
            text="暂无菜品！请前往后台管理添加菜品或检查图片目录。",
            font=("Arial", 11),
        )
        self.notebook.add(empty_label, text="提示")
        return

      for cat in categories:
        cat_frame = ttk.Frame(self.notebook)
        self.notebook.add(cat_frame, text=f"  {cat}  ")

        canvas = tk.Canvas(cat_frame, bg="white", highlightthickness=0)
        scrollbar = ttk.Scrollbar(
            cat_frame, orient="vertical", command=canvas.yview
        )
        scrollable_sub = ttk.Frame(canvas)

        scrollable_sub.bind(
            "<Configure>",
            lambda e, c=canvas: c.configure(scrollregion=c.bbox("all")),
        )
        canvas.create_window((0, 0), window=scrollable_sub, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        cursor.execute(
            "SELECT name, price, image_path FROM dishes WHERE category = ?",
            (cat,),
        )
        dishes = cursor.fetchall()

        row, col = 0, 0
        max_cols = 4
        for dish in dishes:
          name, price, img_path = dish

          card = ttk.LabelFrame(scrollable_sub, text=name, padding=6)
          card.grid(row=row, column=col, padx=8, pady=8, sticky="nsew")

          img_label = ttk.Label(card)
          abs_img_path = resolve_img_path(img_path)
          cache_key = f"{cat}_{name}_{img_path}"

          if abs_img_path and os.path.exists(abs_img_path):
            tk_img = self.get_cached_image(cache_key, abs_img_path, (90, 90))
            img_label.config(image=tk_img)
          else:
            img_label.config(text="[无图片]")
          img_label.pack()

          ttk.Label(card, text=f"¥{price}", font=("Arial", 10, "bold")).pack(
              pady=2
          )
          tk.Button(
              card,
              text="点一份",
              bg="#2196F3",
              fg="white",
              command=lambda n=name, p=price, i=img_path: self.add_to_cart(
                  n, p, i
              ),
          ).pack(fill=tk.X)

          col += 1
          if col >= max_cols:
            col = 0
            row += 1
    finally:
      conn.close()

    self.refresh_order_history()

  def refresh_order_history(self):
    for row in self.order_tree.get_children():
      self.order_tree.delete(row)
    conn = sqlite3.connect(DB_NAME)
    try:
      cursor = conn.cursor()
      cursor.execute(
          "SELECT id, order_time, items_detail, total_amount FROM orders ORDER"
          " BY id DESC"
      )
      for row in cursor.fetchall():
        self.order_tree.insert(
            "",
            tk.END,
            iid=str(row[0]),
            values=(row[1], row[2], f"¥{row[3]:.2f}"),
        )
    finally:
      conn.close()

  def clear_order_history(self):
    if messagebox.askyesno("确认", "确定要清空所有已保存的历史订单记录吗？"):
      conn = sqlite3.connect(DB_NAME)
      try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM orders")
        cursor.execute("DELETE FROM order_items")
        conn.commit()
      finally:
        conn.close()
      self.editing_order_id = None
      self.refresh_order_history()
      self.clear_cart()

  def add_to_cart(self, name, price, image_path):
    if name in self.cart:
      self.cart[name]["count"] += 1
    else:
      self.cart[name] = {
          "price": price,
          "count": 1,
          "image_path": image_path,
      }
    self.update_cart_ui_display()

  def change_count(self, name, delta):
    if name in self.cart:
      self.cart[name]["count"] += delta
      if self.cart[name]["count"] <= 0:
        del self.cart[name]
    self.update_cart_ui_display()

  def remove_from_cart(self, name):
    if name in self.cart:
      del self.cart[name]
    self.update_cart_ui_display()

  def clear_cart(self):
    self.cart.clear()
    for name, ui in list(self.cart_ui_elements.items()):
      ui["frame"].destroy()
    self.cart_ui_elements.clear()
    self.editing_order_id = None
    self.update_cart_ui_display()

  def edit_qty_inline(self, name, btn_frame):
    if name not in self.cart:
      return
    current_count = self.cart[name]["count"]
    for child in btn_frame.winfo_children():
      info = child.grid_info()
      if info.get("column") == 1 and isinstance(child, tk.Label):
        child.destroy()
        break

    entry = tk.Entry(btn_frame, width=2, font=("Arial", 9), justify="center")
    entry.insert(0, str(current_count))
    entry.grid(row=0, column=1, padx=1)
    entry.focus_set()
    entry.selection_range(0, tk.END)

    def save_qty(event=None):
      try:
        new_count = int(entry.get().strip())
        if new_count > 0:
          self.cart[name]["count"] = new_count
        else:
          del self.cart[name]
      except ValueError:
        pass
      self.update_cart_ui_display()

    entry.bind("<Return>", save_qty)
    entry.bind("<FocusOut>", save_qty)

  def update_cart_ui_display(self):
    total_price = 0.0

    existing_names = list(self.cart_ui_elements.keys())
    for name in existing_names:
      if name not in self.cart:
        self.cart_ui_elements[name]["frame"].destroy()
        del self.cart_ui_elements[name]

    row = 0
    for name, info in self.cart.items():
      price = info["price"]
      count = info["count"]
      img_path = info["image_path"]
      subtotal = price * count
      total_price += subtotal

      if name in self.cart_ui_elements:
        ui = self.cart_ui_elements[name]
        ui["frame"].grid(row=row, column=0, sticky="ew", pady=3)
        info_text = f"{name}\n单:¥{price} 小:¥{subtotal:.2f}"
        ui["info_label"].config(text=info_text)
        ui["qty_label"].config(text=str(count))
      else:
        item_frame = ttk.Frame(self.cart_items_frame, padding=2)
        item_frame.grid(row=row, column=0, sticky="ew", pady=3)
        item_frame.columnconfigure(1, weight=1)

        thumb_label = ttk.Label(item_frame)
        abs_img_path = resolve_img_path(img_path)
        cache_key = f"cart_{name}_{img_path}"

        if abs_img_path and os.path.exists(abs_img_path):
          t_img = self.get_cached_image(cache_key, abs_img_path, (32, 32))
          thumb_label.config(image=t_img)
        else:
          thumb_label.config(text="[无]")
        thumb_label.grid(row=0, column=0, padx=2)

        info_text = f"{name}\n单:¥{price} 小:¥{subtotal:.2f}"
        info_label = ttk.Label(item_frame, text=info_text, font=("Arial", 8))
        info_label.grid(row=0, column=1, sticky="w", padx=4)

        btn_frame = ttk.Frame(item_frame)
        btn_frame.grid(row=0, column=2, sticky="e", padx=2)

        tk.Button(
            btn_frame,
            text="-",
            width=1,
            command=lambda n=name: self.change_count(n, -1),
        ).grid(row=0, column=0, padx=1)

        qty_label = tk.Label(
            btn_frame,
            text=str(count),
            font=("Arial", 9),
            bg="white",
            relief="sunken",
            width=2,
            anchor="center",
        )
        qty_label.grid(row=0, column=1, padx=1)
        qty_label.bind(
            "<Double-Button-1>",
            lambda e, n=name, bf=btn_frame: self.edit_qty_inline(n, bf),
        )

        tk.Button(
            btn_frame,
            text="+",
            width=1,
            command=lambda n=name: self.change_count(n, 1),
        ).grid(row=0, column=2, padx=1)

        tk.Button(
            btn_frame,
            text="🗑️",
            command=lambda n=name: self.remove_from_cart(n),
        ).grid(row=0, column=3, padx=1)

        self.cart_ui_elements[name] = {
            "frame": item_frame,
            "info_label": info_label,
            "qty_label": qty_label,
            "btn_frame": btn_frame,
        }

      row += 1

    self.total_label.config(text=f"总计: ¥{total_price:.2f}")

    if self.editing_order_id:
      self.edit_status_label.config(text="⚠️ 正在修改选中历史订单")
      self.cancel_edit_btn.pack(side=tk.LEFT, padx=5)
      self.submit_btn.config(text="保存修改", bg="#FF9800")
    else:
      self.edit_status_label.config(text="")
      self.cancel_edit_btn.pack_forget()
      self.submit_btn.config(text="提交订单", bg="#4CAF50")

  def submit_order(self):
    if not self.cart:
      messagebox.showwarning("提示", "购物车是空的，无法提交！")
      return

    total_price = sum(
        info["price"] * info["count"] for info in self.cart.values()
    )
    items_desc = ", ".join(
        [f"{name}x{info['count']}" for name, info in self.cart.items()]
    )
    order_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn = sqlite3.connect(DB_NAME)
    try:
      cursor = conn.cursor()

      if self.editing_order_id:
        cursor.execute(
            """
                    UPDATE orders 
                    SET items_detail = ?, total_amount = ?, order_time = ? 
                    WHERE id = ?
                """,
            (items_desc, total_price, order_time, self.editing_order_id),
        )
        cursor.execute(
            "DELETE FROM order_items WHERE order_id = ?",
            (self.editing_order_id,),
        )
        order_id = self.editing_order_id
        success_msg = "历史订单修改保存成功！"
      else:
        cursor.execute(
            """
                    INSERT INTO orders (order_time, items_detail, total_amount) 
                    VALUES (?, ?, ?)
                """,
            (order_time, items_desc, total_price),
        )
        order_id = cursor.lastrowid
        success_msg = f"新订单提交成功！总计 ¥{total_price:.2f}"

      for name, info in self.cart.items():
        cursor.execute(
            """
                    INSERT INTO order_items (order_id, name, price, count, image_path) 
                    VALUES (?, ?, ?, ?, ?)
                """,
            (
                order_id,
                name,
                info["price"],
                info["count"],
                info["image_path"],
            ),
        )

      conn.commit()
    finally:
      conn.close()

    messagebox.showinfo("成功", success_msg)
    self.editing_order_id = None
    self.clear_cart()
    self.refresh_order_history()


# --- 第二页：后台管理中心 ---
class PageAdmin(tk.Frame):

  def __init__(self, parent, page1_ref):
    super().__init__(parent)
    self.grid(row=0, column=0, sticky="nsew")
    self.page1_ref = page1_ref
    self.image_cache = {}

    self.admin_notebook = ttk.Notebook(self)
    self.admin_notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    self.tab_dishes = ttk.Frame(self.admin_notebook)
    self.tab_transactions = ttk.Frame(self.admin_notebook)

    self.admin_notebook.add(self.tab_dishes, text="  🍜 菜品管理  ")
    self.admin_notebook.add(self.tab_transactions, text="  📊 交易流水与智能订单生成  ")

    self.setup_dish_management_ui()
    self.setup_transaction_ui()

  def refresh_data(self):
    self.refresh_dish_tree()
    self.refresh_transaction_tree()

  def setup_dish_management_ui(self):
    dish_layout = ttk.Frame(self.tab_dishes, padding=10)
    dish_layout.pack(fill=tk.BOTH, expand=True)

    tree_frame = tk.Frame(dish_layout)
    tree_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    self.dish_tree = ttk.Treeview(
        tree_frame,
        columns=("ID", "菜品名称", "价格", "分类", "图片路径"),
        show="headings",
    )
    self.dish_tree.heading("ID", text="ID")
    self.dish_tree.heading("菜品名称", text="菜品名称")
    self.dish_tree.heading("价格", text="价格 (¥)")
    self.dish_tree.heading("分类", text="分类")
    self.dish_tree.heading("图片路径", text="图片路径 (相对路径)")

    self.dish_tree.column("ID", width=40, anchor="center")
    self.dish_tree.column("菜品名称", width=120, anchor="w")
    self.dish_tree.column("价格", width=80, anchor="center")
    self.dish_tree.column("分类", width=100, anchor="center")
    self.dish_tree.column("图片路径", width=300, anchor="w")

    self.dish_tree.bind("<<TreeviewSelect>>", self.on_dish_select)

    scrollbar = ttk.Scrollbar(
        tree_frame, orient="vertical", command=self.dish_tree.yview
    )
    self.dish_tree.configure(yscrollcommand=scrollbar.set)
    self.dish_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    form_frame = ttk.LabelFrame(
        dish_layout, text=" 📝 菜品信息编辑 ", width=350, padding=15
    )
    form_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=False, padx=(10, 0))
    form_frame.pack_propagate(False)

    ttk.Label(form_frame, text="菜品名称:").pack(anchor="w", pady=2)
    self.dish_name_entry = ttk.Entry(form_frame, font=("Arial", 10))
    self.dish_name_entry.pack(fill=tk.X, pady=2)

    ttk.Label(form_frame, text="价格 (¥):").pack(anchor="w", pady=2)
    self.dish_price_entry = ttk.Entry(form_frame, font=("Arial", 10))
    self.dish_price_entry.pack(fill=tk.X, pady=2)

    ttk.Label(form_frame, text="分类:").pack(anchor="w", pady=2)
    self.dish_cat_entry = ttk.Entry(form_frame, font=("Arial", 10))
    self.dish_cat_entry.pack(fill=tk.X, pady=2)

    ttk.Label(form_frame, text="图片路径:").pack(anchor="w", pady=2)
    path_box = tk.Frame(form_frame)
    path_box.pack(fill=tk.X, pady=2)
    self.dish_img_entry = ttk.Entry(path_box, font=("Arial", 9))
    self.dish_img_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
    tk.Button(
        path_box, text="选择...", command=self.browse_dish_image, font=("Arial", 9)
    ).pack(side=tk.RIGHT, padx=2)

    btn_box = tk.Frame(form_frame)
    btn_box.pack(fill=tk.X, pady=20)

    tk.Button(
        btn_box,
        text="添加新菜品",
        command=self.add_dish,
        bg="#4CAF50",
        fg="white",
        font=("Arial", 10),
    ).pack(fill=tk.X, pady=3)
    tk.Button(
        btn_box,
        text="保存修改",
        command=self.update_dish,
        bg="#2196F3",
        fg="white",
        font=("Arial", 10),
    ).pack(fill=tk.X, pady=3)
    tk.Button(
        btn_box,
        text="删除选中菜品",
        command=self.delete_dish,
        bg="#f44336",
        fg="white",
        font=("Arial", 10),
    ).pack(fill=tk.X, pady=3)
    tk.Button(
        btn_box,
        text="清空输入",
        command=self.clear_dish_form,
        bg="#e0e0e0",
        font=("Arial", 10),
    ).pack(fill=tk.X, pady=3)

    self.selected_dish_id = None

  def refresh_dish_tree(self):
    for row in self.dish_tree.get_children():
      self.dish_tree.delete(row)
    conn = sqlite3.connect(DB_NAME)
    try:
      cursor = conn.cursor()
      cursor.execute("SELECT id, name, price, category, image_path FROM dishes")
      for row in cursor.fetchall():
        self.dish_tree.insert("", tk.END, values=row)
    finally:
      conn.close()

  def on_dish_select(self, event):
    selected = self.dish_tree.selection()
    if not selected:
      return
    values = self.dish_tree.item(selected[0], "values")
    if not values:
      return
    self.selected_dish_id = values[0]
    self.dish_name_entry.delete(0, tk.END)
    self.dish_name_entry.insert(0, values[1])
    self.dish_price_entry.delete(0, tk.END)
    self.dish_price_entry.insert(0, values[2])
    self.dish_cat_entry.delete(0, tk.END)
    self.dish_cat_entry.insert(0, values[3])
    self.dish_img_entry.delete(0, tk.END)
    self.dish_img_entry.insert(0, values[4])

  def browse_dish_image(self):
    f_path = filedialog.askopenfilename(
        filetypes=[("Image files", "*.png *.jpg *.jpeg *.webp")]
    )
    if f_path:
      category = self.dish_cat_entry.get().strip() or "默认分类"
      target_folder = os.path.join(IMAGE_DIR, category)
      os.makedirs(target_folder, exist_ok=True)

      filename = os.path.basename(f_path)
      dest_path = os.path.join(target_folder, filename)

      try:
        if os.path.abspath(f_path) != os.path.abspath(dest_path):
          shutil.copy(f_path, dest_path)
        rel_path = get_relative_path(dest_path)
        self.dish_img_entry.delete(0, tk.END)
        self.dish_img_entry.insert(0, rel_path)
      except Exception as e:
        messagebox.showerror("错误", f"复制图片失败: {e}")

  def clear_dish_form(self):
    self.selected_dish_id = None
    self.dish_name_entry.delete(0, tk.END)
    self.dish_price_entry.delete(0, tk.END)
    self.dish_cat_entry.delete(0, tk.END)
    self.dish_img_entry.delete(0, tk.END)

  def add_dish(self):
    name = self.dish_name_entry.get().strip()
    try:
      price = float(self.dish_price_entry.get().strip())
    except ValueError:
      messagebox.showerror("错误", "请输入有效的价格数字！")
      return
    category = self.dish_cat_entry.get().strip() or "默认分类"
    img_path = get_relative_path(self.dish_img_entry.get().strip())

    if not name:
      messagebox.showwarning("提示", "菜品名称不能为空！")
      return

    conn = sqlite3.connect(DB_NAME)
    try:
      cursor = conn.cursor()
      cursor.execute(
          "INSERT INTO dishes (name, price, category, image_path) VALUES (?, ?,"
          " ?, ?)",
          (name, price, category, img_path),
      )
      conn.commit()
    finally:
      conn.close()

    self.refresh_dish_tree()
    self.clear_dish_form()
    self.page1_ref.is_initialized = False
    self.page1_ref.pil_image_cache.clear()
    messagebox.showinfo("成功", "新菜品添加成功！")

  def update_dish(self):
    if not self.selected_dish_id:
      messagebox.showwarning("提示", "请先在左侧列表中选中要修改的菜品！")
      return
    name = self.dish_name_entry.get().strip()
    try:
      price = float(self.dish_price_entry.get().strip())
    except ValueError:
      messagebox.showerror("错误", "请输入有效的价格数字！")
      return
    category = self.dish_cat_entry.get().strip() or "默认分类"
    img_path = get_relative_path(self.dish_img_entry.get().strip())

    conn = sqlite3.connect(DB_NAME)
    try:
      cursor = conn.cursor()
      cursor.execute(
          """
                UPDATE dishes SET name = ?, price = ?, category = ?, image_path = ? WHERE id = ?
            """,
          (name, price, category, img_path, self.selected_dish_id),
      )
      conn.commit()
    finally:
      conn.close()

    self.refresh_dish_tree()
    self.page1_ref.is_initialized = False
    self.page1_ref.pil_image_cache.clear()
    messagebox.showinfo("成功", f"菜品 ID #{self.selected_dish_id} 修改成功！")

  def delete_dish(self):
    if not self.selected_dish_id:
      messagebox.showwarning("提示", "请先在左侧列表中选中要删除的菜品！")
      return
    if messagebox.askyesno(
        "确认删除", f"确定要删除菜品 ID #{self.selected_dish_id} 吗？"
    ):
      conn = sqlite3.connect(DB_NAME)
      try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM dishes WHERE id = ?", (self.selected_dish_id,))
        conn.commit()
      finally:
        conn.close()

      self.refresh_dish_tree()
      self.clear_dish_form()
      self.page1_ref.is_initialized = False
      self.page1_ref.pil_image_cache.clear()
      messagebox.showinfo("成功", "菜品删除成功！")

  def setup_transaction_ui(self):
    trans_layout = ttk.Frame(self.tab_transactions, padding=10)
    trans_layout.pack(fill=tk.BOTH, expand=True)

    left_container = tk.Frame(trans_layout)
    left_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))

    top_bar = tk.Frame(left_container)
    top_bar.pack(side=tk.TOP, fill=tk.X, pady=5)

    tk.Button(
        top_bar,
        text="📂 导入交易流水 Excel 表格",
        command=self.import_transactions_excel,
        bg="#2196F3",
        fg="white",
        font=("Arial", 10, "bold"),
    ).pack(side=tk.LEFT, padx=2)

    tree_frame = tk.Frame(left_container)
    tree_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, pady=5)

    self.trans_tree = ttk.Treeview(
        tree_frame, columns=("交易时间", "交易金额"), show="headings"
    )
    self.trans_tree.heading("交易时间", text="交易时间")
    self.trans_tree.heading("交易金额", text="交易金额 (元)")

    self.trans_tree.column("交易时间", width=180, anchor="center")
    self.trans_tree.column("交易金额", width=120, anchor="center")

    self.trans_tree.bind("<<TreeviewSelect>>", self.on_transaction_select)

    trans_scrollbar = ttk.Scrollbar(
        tree_frame, orient="vertical", command=self.trans_tree.yview
    )
    self.trans_tree.configure(yscrollcommand=trans_scrollbar.set)

    self.trans_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    trans_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    right_container = ttk.LabelFrame(
        trans_layout, text=" ⚙️ 智能逻辑条件与订单生成 ", width=420, padding=10
    )
    right_container.pack(side=tk.RIGHT, fill=tk.BOTH, expand=False, padx=(5, 0))
    right_container.pack_propagate(False)

    ttk.Label(
        right_container,
        text="请输入逻辑条件规则 (例如: 必须含酒水; 大额优先硬菜):",
        font=("Arial", 9, "bold"),
    ).pack(anchor="w", pady=(2, 2))
    self.rule_condition_entry = ttk.Entry(right_container, font=("Arial", 9))
    self.rule_condition_entry.pack(fill=tk.X, pady=2)
    self.rule_condition_entry.insert(
        0, "条件: 强制含酒水饮料; 按金额阶梯匹配硬菜与小炒"
    )

    tk.Button(
        right_container,
        text="🔄 按条件逻辑生成随机订单",
        command=self.generate_random_orders_from_transactions,
        bg="#FF9800",
        fg="white",
        font=("Arial", 10, "bold"),
    ).pack(side=tk.TOP, fill=tk.X, pady=8)

    ttk.Label(
        right_container,
        text="选中流事的对应订单图文详情凭证：",
        font=("Arial", 9, "bold"),
    ).pack(anchor="w", pady=(10, 2))

    detail_container = tk.Frame(right_container, bg="white")
    detail_container.pack(side=tk.TOP, fill=tk.BOTH, expand=True, pady=2)

    self.detail_canvas = tk.Canvas(
        detail_container, bg="white", highlightthickness=0
    )
    detail_scrollbar = ttk.Scrollbar(
        detail_container, orient="vertical", command=self.detail_canvas.yview
    )
    self.detail_items_frame = ttk.Frame(self.detail_canvas)

    self.detail_items_frame.bind(
        "<Configure>",
        lambda e: self.detail_canvas.configure(
            scrollregion=self.detail_canvas.bbox("all")
        ),
    )
    self.detail_canvas.create_window(
        (0, 0), window=self.detail_items_frame, anchor="nw"
    )
    self.detail_canvas.configure(yscrollcommand=detail_scrollbar.set)

    self.detail_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    detail_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    export_box = tk.Frame(right_container)
    export_box.pack(side=tk.BOTTOM, fill=tk.X, pady=5)

    tk.Button(
        export_box,
        text="📥 导出表格",
        command=self.export_orders_excel,
        bg="#4CAF50",
        fg="white",
        font=("Arial", 9),
    ).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)

    tk.Button(
        export_box,
        text="🖼️ 生成订单图片",
        command=self.generate_order_images_files,
        bg="#9C27B0",
        fg="white",
        font=("Arial", 9, "bold"),
    ).pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=2)

    self.selected_trans_id = None

  def refresh_transaction_tree(self):
    for row in self.trans_tree.get_children():
      self.trans_tree.delete(row)
    conn = sqlite3.connect(DB_NAME)
    try:
      cursor = conn.cursor()
      cursor.execute("SELECT id, trans_time, amount FROM transactions")
      rows = cursor.fetchall()
      for row in rows:
        self.trans_tree.insert(
            "", tk.END, iid=str(row[0]), values=(row[1], row[2])
        )
    finally:
      conn.close()

  def import_transactions_excel(self):
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if not file_path:
      return

    try:
      wb = openpyxl.load_workbook(file_path)
      sheet = wb.active
      rows = list(sheet.iter_rows(values_only=True))
      if not rows:
        messagebox.showerror("错误", "Excel 文件内容为空！")
        return

      header_row = rows[0]
      time_idx, amount_idx = -1, -1
      for idx, val in enumerate(header_row):
        if val is not None:
          h_str = str(val).strip()
          if "交易时间" in h_str:
            time_idx = idx
          elif "交易金额" in h_str:
            amount_idx = idx

      if time_idx == -1:
        time_idx = 0
      if amount_idx == -1 and len(header_row) > 1:
        amount_idx = 1

      conn = sqlite3.connect(DB_NAME)
      try:
        cursor = conn.cursor()
        count = 0
        for row in rows[1:]:
          if len(row) > max(time_idx, amount_idx):
            t_val = row[time_idx]
            a_val = row[amount_idx]
            if t_val is not None and a_val is not None:
              trans_time = str(t_val).strip()
              try:
                clean_a = str(a_val).replace("¥", "").replace(",", "").strip()
                amount = float(clean_a)
              except (ValueError, TypeError):
                amount = 0.0

              cursor.execute(
                  "INSERT INTO transactions (trans_time, amount) VALUES (?, ?)",
                  (trans_time, amount),
              )
              count += 1
        conn.commit()
      finally:
        conn.close()

      self.refresh_transaction_tree()
      messagebox.showinfo("成功", f"成功导入 {count} 条交易流水记录！")
    except Exception as e:
      messagebox.showerror("错误", f"导入失败: {e}")

  def on_transaction_select(self, event):
    selected = self.trans_tree.selection()
    if not selected:
      return
    self.selected_trans_id = selected[0]
    values = self.trans_tree.item(selected[0], "values")
    if not values:
      return

    trans_time = values[0]
    trans_amount = values[1]

    for widget in self.detail_items_frame.winfo_children():
      widget.destroy()

    conn = sqlite3.connect(DB_NAME)
    try:
      cursor = conn.cursor()
      cursor.execute(
          "SELECT id, order_time, items_detail, total_amount FROM orders WHERE"
          " trans_id = ?",
          (self.selected_trans_id,),
      )
      order_row = cursor.fetchone()

      if order_row:
        o_id, o_time, items_detail, total_amt = order_row
        cursor.execute(
            "SELECT name, price, count, image_path FROM order_items WHERE"
            " order_id = ?",
            (o_id,),
        )
        items = cursor.fetchall()

        img_width = 380
        img_height = 75 + len(items) * 58 + 55

        canvas_img = Image.new("RGB", (img_width, img_height), (255, 255, 255))
        draw = ImageDraw.Draw(canvas_img)

        font_title = get_pil_font(13)
        font_text = get_pil_font(11)
        font_small = get_pil_font(9)

        draw.text(
            (15, 12), "餐厅消费订单凭证", fill=(33, 33, 33), font=font_title
        )
        draw.text(
            (15, 32),
            f"下单时间: {o_time}",
            fill=(100, 100, 100),
            font=font_small,
        )
        draw.line(
            [(15, 48), (img_width - 15, 48)], fill=(220, 220, 220), width=1
        )

        current_y = 56
        for item in items:
          name, price, count, img_path = item
          subtotal = price * count

          thumb_box = [15, current_y, 59, current_y + 44]
          draw.rectangle(thumb_box, outline=(230, 230, 230), width=1)

          abs_img_path = resolve_img_path(img_path)
          if abs_img_path and os.path.exists(abs_img_path):
            try:
              p_img = Image.open(abs_img_path).convert("RGB")
              p_img.thumbnail((36, 36))
              pw, ph = p_img.size
              px = 15 + (44 - pw) // 2
              py = current_y + (44 - ph) // 2
              canvas_img.paste(p_img, (px, py))
            except Exception:
              draw.text(
                  (20, current_y + 14),
                  "无图",
                  fill=(150, 150, 150),
                  font=font_small,
              )
          else:
            draw.text(
                (20, current_y + 14),
                "无图",
                fill=(150, 150, 150),
                font=font_small,
            )

          draw.text(
              (69, current_y + 3), str(name), fill=(33, 33, 33), font=font_text
          )
          draw.text(
              (69, current_y + 22),
              f"单价: ¥{price:.2f}  x  {count}",
              fill=(120, 120, 120),
              font=font_small,
          )

          sub_str = f"¥{subtotal:.2f}"
          draw.text(
              (img_width - 15 - len(sub_str) * 7, current_y + 12),
              sub_str,
              fill=(211, 47, 47),
              font=font_text,
          )

          current_y += 58

        draw.line(
            [(15, current_y + 4), (img_width - 15, current_y + 4)],
            fill=(220, 220, 220),
            width=1,
        )
        draw.text(
            (15, current_y + 14), "订单总计:", fill=(33, 33, 33), font=font_title
        )
        total_str = f"¥{total_amt:.2f}"
        draw.text(
            (img_width - 15 - len(total_str) * 9, current_y + 14),
            total_str,
            fill=(211, 47, 47),
            font=font_title,
        )

        tk_preview_img = ImageTk.PhotoImage(canvas_img)
        self.image_cache[f"preview_{o_id}"] = tk_preview_img

        preview_label = tk.Label(
            self.detail_items_frame,
            image=tk_preview_img,
            bg="white",
            relief="solid",
            borderwidth=1,
        )
        preview_label.pack(anchor="w", padx=2, pady=5)
      else:
        tk.Label(
            self.detail_items_frame,
            text="\n【提示】该流水尚未生成随机订单。",
            fg="gray",
            bg="white",
            font=("Arial", 9),
        ).pack(anchor="w", pady=10)
    finally:
      conn.close()

  def generate_random_orders_from_transactions(self):
    condition_rule = self.rule_condition_entry.get().strip()
    conn = sqlite3.connect(DB_NAME)
    try:
      cursor = conn.cursor()

      cursor.execute("SELECT id, name, price, category, image_path FROM dishes")
      all_dishes = cursor.fetchall()

      if not all_dishes:
        messagebox.showerror(
            "错误", "菜品库为空！请先前往菜品管理添加菜品或检查图片目录。"
        )
        return

      cursor.execute("SELECT id, trans_time, amount FROM transactions")
      transactions = cursor.fetchall()

      if not transactions:
        messagebox.showwarning("提示", "请先导入交易流水表格再点击生成！")
        return

      cursor.execute(
          "DELETE FROM order_items WHERE order_id IN (SELECT id FROM orders"
          " WHERE trans_id IS NOT NULL)"
      )
      cursor.execute("DELETE FROM orders WHERE trans_id IS NOT NULL")

      def get_dish_category_type(cat_name, dish_name):
        c_lower = (cat_name or "").lower()
        n_lower = (dish_name or "").lower()
        if (
            "酒" in c_lower
            or "饮" in c_lower
            or "水" in c_lower
            or "酒" in n_lower
            or "饮" in n_lower
            or "可乐" in n_lower
            or "雪碧" in n_lower
            or "啤酒" in n_lower
            or "果汁" in n_lower
            or "茶" in n_lower
        ):
          return "酒水饮料"
        elif (
            "主食" in c_lower
            or "饭" in n_lower
            or "面" in n_lower
            or "粉" in n_lower
            or "饺" in n_lower
            or "饼" in n_lower
        ):
          return "主食"
        elif "汤" in c_lower or "汤" in n_lower:
          return "汤类"
        elif "冷" in c_lower or "凉" in c_lower or "凉菜" in n_lower:
          return "冷菜"
        elif "小吃" in c_lower or "甜品" in c_lower or "点心" in n_lower:
          return "小吃点心"
        else:
          return "热菜"

      categories_dict = {
          "热菜": [],
          "冷菜": [],
          "主食": [],
          "汤类": [],
          "酒水饮料": [],
          "小吃点心": [],
      }

      for d in all_dishes:
        d_id, d_name, d_price, d_cat, d_img = d
        c_type = get_dish_category_type(d_cat, d_name)
        categories_dict[c_type].append(d)

      drink_pool = [d for d in categories_dict["酒水饮料"] if d[2] > 0]
      if not drink_pool:
        drink_pool = [d for d in all_dishes if d[2] > 0]
        if not drink_pool:
          drink_pool = all_dishes

      generated_count = 0

      for trans in transactions:
        trans_id, trans_time, target_amount = trans
        target_amount = float(target_amount)

        order_cart = {}
        remaining = target_amount

        if "酒" in condition_rule or "含" in condition_rule:
          valid_drinks = [d for d in drink_pool if 0 < d[2] <= remaining]
          if not valid_drinks:
            valid_drinks = drink_pool
          if valid_drinks:
            chosen_drink = random.choice(valid_drinks)
            d_id, d_name, d_price, d_cat, d_img = chosen_drink
            if d_price <= 0:
              d_price = 3.0
            d_count = 1
            if remaining < d_price:
              d_price = max(1.0, remaining)

            order_cart[d_name] = {
                "price": d_price,
                "count": d_count,
                "image_path": d_img,
            }
            remaining -= d_price * d_count

        active_categories = [
            cat for cat, items in categories_dict.items() if items
        ]
        if not active_categories:
          active_categories = ["热菜"]

        while remaining > 0:
          chosen_cat = random.choice(active_categories)
          cat_dishes = [d for d in categories_dict[chosen_cat] if d[2] > 0]
          if not cat_dishes:
            cat_dishes = [d for d in all_dishes if d[2] > 0]
            if not cat_dishes:
              cat_dishes = all_dishes

          chosen_dish = random.choice(cat_dishes)
          c_id, c_name, c_price, c_cat, c_img = chosen_dish

          if c_price <= 0:
            c_price = 1.0

          if remaining >= c_price:
            max_c = int(remaining // c_price)
            count = random.randint(1, max(1, min(max_c, 2)))
          else:
            count = 1
            c_price = remaining

          sub_cost = c_price * count
          if c_name in order_cart:
            order_cart[c_name]["count"] += count
          else:
            order_cart[c_name] = {
                "price": c_price,
                "count": count,
                "image_path": c_img,
            }

          remaining -= sub_cost
          if remaining < 0.1:
            break

        current_total = sum(
            info["price"] * info["count"] for info in order_cart.values()
        )
        if order_cart and abs(current_total - target_amount) > 0.001:
          first_key = list(order_cart.keys())[0]
          diff = target_amount - current_total
          if order_cart[first_key]["count"] > 0:
            order_cart[first_key]["price"] += (
                diff / order_cart[first_key]["count"]
            )
            if order_cart[first_key]["price"] < 0:
              order_cart[first_key]["price"] = 1.0

        items_desc = ", ".join(
            [f"{name}x{info['count']}" for name, info in order_cart.items()]
        )
        actual_total = sum(
            info["price"] * info["count"] for info in order_cart.values()
        )

        cursor.execute(
            """
                    INSERT INTO orders (order_time, items_detail, total_amount, trans_id) 
                    VALUES (?, ?, ?, ?)
                """,
            (trans_time, items_desc, actual_total, trans_id),
        )
        new_order_id = cursor.lastrowid

        for name, info in order_cart.items():
          cursor.execute(
              """
                        INSERT INTO order_items (order_id, name, price, count, image_path) 
                        VALUES (?, ?, ?, ?, ?)
                    """,
              (
                  new_order_id,
                  name,
                  info["price"],
                  info["count"],
                  info["image_path"],
              ),
          )
        generated_count += 1

      conn.commit()
    finally:
      conn.close()

    self.refresh_transaction_tree()
    self.page1_ref.refresh_order_history()
    messagebox.showinfo(
        "成功", f"成功为 {generated_count} 条交易流水生成匹配订单！"
    )

  def export_orders_excel(self):
    # 弹出设置窗口，自定义时间跨度统计（默认每天）
    dialog = tk.Toplevel(self)
    dialog.title("导出与统计设置")
    dialog.geometry("320x210")
    dialog.resizable(False, False)
    dialog.grab_set()

    ttk.Label(
        dialog,
        text="请选择销售统计的时间跨度：",
        font=("Arial", 10, "bold"),
    ).pack(pady=12)

    span_var = tk.StringVar(value="daily")

    ttk.Radiobutton(
        dialog, text="每天 (Daily) —— 默认", variable=span_var, value="daily"
    ).pack(anchor="w", padx=50, pady=3)
    ttk.Radiobutton(
        dialog, text="每周 (Weekly)", variable=span_var, value="weekly"
    ).pack(anchor="w", padx=50, pady=3)
    ttk.Radiobutton(
        dialog, text="每月 (Monthly)", variable=span_var, value="monthly"
    ).pack(anchor="w", padx=50, pady=3)

    def do_export():
      span = span_var.get()
      dialog.destroy()

      file_path = filedialog.asksaveasfilename(
          defaultextension=".xlsx",
          filetypes=[("Excel files", "*.xlsx")],
          initialfile="餐厅订单明细及销售统计表.xlsx",
      )
      if not file_path:
        return

      try:
        wb = openpyxl.Workbook()

        # Sheet 1: 订单明细（4列：订单编号、订单时间、订单金额、订单详情）
        ws1 = wb.active
        ws1.title = "订单明细"
        ws1.append(["订单编号", "订单时间", "订单金额 (元)", "订单详情"])

        conn = sqlite3.connect(DB_NAME)
        try:
          cursor = conn.cursor()
          # 严格按照时间正序排列并编号
          cursor.execute(
              "SELECT order_time, total_amount, items_detail FROM orders ORDER"
              " BY order_time ASC"
          )
          rows = cursor.fetchall()

          order_records = []
          daily_counter = {}  # 用于记录每天的订单序号
          for row in rows:
            o_time, o_amt, o_items = row

            # 解析订单时间，提取年月日
            try:
              dt = datetime.strptime(o_time.strip(), "%Y-%m-%d %H:%M:%S")
            except ValueError:
              try:
                dt = datetime.strptime(o_time.strip()[:10], "%Y-%m-%d")
              except Exception:
                dt = datetime.now()

            date_prefix = dt.strftime("%Y%m%d")
            if date_prefix not in daily_counter:
              daily_counter[date_prefix] = 1
            else:
              daily_counter[date_prefix] += 1

            # 生成格式如 202608080001 的编号
            order_no = f"{date_prefix}{daily_counter[date_prefix]:04d}"

            ws1.append([order_no, o_time, o_amt, o_items])
            order_records.append((o_time, o_amt))

          # Sheet 2: 销售统计
          ws2 = wb.create_sheet(title="销售统计")
          ws2.append(["统计时间段", "订单总数", "销售总额 (元)", "平均客单价 (元)"])

          stats_dict = {}
          for o_time, o_amt in order_records:
            try:
              dt = datetime.strptime(o_time.strip(), "%Y-%m-%d %H:%M:%S")
            except ValueError:
              try:
                dt = datetime.strptime(o_time.strip()[:10], "%Y-%m-%d")
              except Exception:
                dt = datetime.now()

            if span == "daily":
              key = dt.strftime("%Y-%m-%d")
            elif span == "weekly":
              y, w, _ = dt.isocalendar()
              key = f"{y}-W{w:02d}"
            elif span == "monthly":
              key = dt.strftime("%Y-%m")
            else:
              key = dt.strftime("%Y-%m-%d")

            if key not in stats_dict:
              stats_dict[key] = {"count": 0, "total": 0.0}
            stats_dict[key]["count"] += 1
            stats_dict[key]["total"] += float(o_amt)

          for key in sorted(stats_dict.keys()):
            count = stats_dict[key]["count"]
            total = stats_dict[key]["total"]
            avg = total / count if count > 0 else 0.0
            ws2.append([key, count, round(total, 2), round(avg, 2)])

          wb.save(file_path)
        finally:
          conn.close()

        messagebox.showinfo("成功", f"订单明细及统计表已成功导出至:\n{file_path}")
      except Exception as e:
        messagebox.showerror("错误", f"导出 Excel 失败: {e}")

    tk.Button(
        dialog,
        text="确认导出",
        command=do_export,
        bg="#4CAF50",
        fg="white",
        font=("Arial", 9, "bold"),
        width=15,
    ).pack(pady=12)

  def generate_order_images_files(self):
    dir_path = filedialog.askdirectory(title="选择保存订单图片的文件夹")
    if not dir_path:
      return

    try:
      conn = sqlite3.connect(DB_NAME)
      try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, order_time, total_amount FROM orders WHERE trans_id IS"
            " NOT NULL"
        )
        orders = cursor.fetchall()

        if not orders:
          messagebox.showwarning("提示", "当前没有生成的随机订单数据可供生成图片！")
          return

        font_title = get_pil_font(14)
        font_text = get_pil_font(12)
        font_small = get_pil_font(10)

        success_count = 0
        for idx, order in enumerate(orders, 1):
          order_id, order_time, total_amount = order

          cursor.execute(
              "SELECT name, price, count, image_path FROM order_items WHERE"
              " order_id = ?",
              (order_id,),
          )
          items = cursor.fetchall()

          img_width = 420
          img_height = 70 + len(items) * 64 + 60

          canvas = Image.new("RGB", (img_width, img_height), (255, 255, 255))
          draw = ImageDraw.Draw(canvas)

          draw.text((20, 15), f"餐厅消费订单凭证", fill=(33, 33, 33), font=font_title)
          draw.text(
              (20, 38),
              f"下单时间: {order_time}",
              fill=(100, 100, 100),
              font=font_small,
          )
          draw.line(
              [(20, 55), (img_width - 20, 55)], fill=(220, 220, 220), width=1
          )

          current_y = 65
          for item in items:
            name, price, count, img_path = item
            subtotal = price * count

            thumb_box = [20, current_y, 72, current_y + 52]
            draw.rectangle(thumb_box, outline=(230, 230, 230), width=1)

            abs_img_path = resolve_img_path(img_path)
            if abs_img_path and os.path.exists(abs_img_path):
              try:
                p_img = Image.open(abs_img_path).convert("RGB")
                p_img.thumbnail((48, 48))
                pw, ph = p_img.size
                px = 20 + (52 - pw) // 2
                py = current_y + (52 - ph) // 2
                canvas.paste(p_img, (px, py))
              except Exception:
                draw.text(
                    (26, current_y + 18),
                    "无图",
                    fill=(150, 150, 150),
                    font=font_small,
                )
            else:
              draw.text(
                  (26, current_y + 18),
                  "无图",
                  fill=(150, 150, 150),
                  font=font_small,
              )

            draw.text(
                (85, current_y + 6), str(name), fill=(33, 33, 33), font=font_text
            )
            draw.text(
                (85, current_y + 28),
                f"单价: ¥{price:.2f}  x  {count}",
                fill=(120, 120, 120),
                font=font_small,
            )

            draw.text(
                (img_width - 25 - len(f"¥{subtotal:.2f}") * 7, current_y + 16),
                f"¥{subtotal:.2f}",
                fill=(211, 47, 47),
                font=font_text,
            )

            current_y += 64

          draw.line(
              [(20, current_y + 5), (img_width - 20, current_y + 5)],
              fill=(220, 220, 220),
              width=1,
          )
          draw.text(
              (20, current_y + 18),
              "订单总计:",
              fill=(33, 33, 33),
              font=font_title,
          )
          total_str = f"¥{total_amount:.2f}"
          draw.text(
              (img_width - 20 - len(total_str) * 10, current_y + 16),
              total_str,
              fill=(211, 47, 47),
              font=font_text,
          )

          out_file = os.path.join(dir_path, f"订单凭证_{idx}.png")
          canvas.save(out_file)
          success_count += 1
      finally:
        conn.close()

      messagebox.showinfo("成功", f"成功为 {success_count} 个订单生成了图片！")
    except Exception as e:
      messagebox.showerror("错误", f"生成订单图片失败: {e}")


if __name__ == "__main__":
  app = RestaurantApp()
  app.mainloop()
